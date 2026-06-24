"""Build a donor/demand :class:`~steelreuse.schema.ExtractedModel` from a spreadsheet.

The extractor writes JSON, but a stockist, demolition contractor or engineer who has *no* Revit
model still has a list of members — usually in Excel or CSV. This module is the bridge: it emits a
**blank inventory template** (headers + one worked example row, including the conservative
``unverified`` provenance flag) and reads a filled-in spreadsheet (``.xlsx`` / ``.csv`` / ``.json``)
back into the same schema the rest of the pipeline already consumes.

Pure (stdlib only, except the optional ``openpyxl`` path for ``.xlsx``), so it is fully unit-tested
and runs in the headless engine. It deliberately reuses :mod:`steelreuse.survey` for the row reading,
header canonicalisation and audit-value normalisation rather than duplicating that machinery — this
is the *inventory* sibling of the *survey* importer: survey augments an existing model with audit
data keyed by element, this one builds the model itself.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from .schema import ROLES, ExtractedMember, ExtractedModel
from .survey import _canon_header, _read_rows, normalize_survey_value

# The columns a person fills in to describe a reclaimed-steel inventory, in template order. The first
# group identifies and sizes the member (what the matcher needs); the second is the pre-demolition
# audit / provenance block (optional, but the honest default for un-surveyed stock is ``unverified``).
_IDENTITY_COLUMNS = ["id", "mark", "unique_id", "role", "level", "raw_section", "material_grade",
                     "length_mm", "recoverable_length_mm"]
_AUDIT_COLUMNS = ["condition_grade", "verification_status", "knockdown", "defects",
                  "connection_type", "connection_condition", "deconstructability"]
INVENTORY_COLUMNS = _IDENTITY_COLUMNS + _AUDIT_COLUMNS

# A single worked example row, so the template documents itself. The ``unverified`` flag is the point:
# stock with no mill cert / coupon test is admitted only under --include-unverified (otherwise
# quarantined), at a conservative knockdown — the template shows that default rather than hiding it.
INVENTORY_EXAMPLE = {
    "id": "D1",
    "mark": "B-12",
    "unique_id": "",
    "role": "beam",
    "level": "L02",
    "raw_section": "IPE 300",
    "material_grade": "S275",
    "length_mm": 8000,
    "recoverable_length_mm": 7600,
    "condition_grade": "B",
    "verification_status": "unverified",
    "knockdown": "",
    "defects": "light surface corrosion",
    "connection_type": "bolted",
    "connection_condition": "B",
    "deconstructability": "easy",
}

# One-line help per column, for the Excel template's "Guide" sheet.
COLUMN_HELP = {
    "id": "REQUIRED. Any stable identifier for the member (piece number, tag, row label).",
    "mark": "Optional human mark, e.g. B-12. Used as the id if id is blank.",
    "unique_id": "Optional Revit UniqueId / IFC GlobalId; leave blank for a hand-built inventory.",
    "role": "beam | column | brace (anything else -> unknown).",
    "level": "Optional storey/level label.",
    "raw_section": "Section designation as written, e.g. 'IPE 300', 'HE 300 A', 'UB 305x165x40'. "
                   "Matched to the catalogue automatically.",
    "material_grade": "Steel grade, e.g. S235 / S275 / S355. Blank -> the run's grade default.",
    "length_mm": "REQUIRED. Physical member (stock) length in millimetres.",
    "recoverable_length_mm": "Usable length after deconstruction (mm). Blank -> equals length_mm.",
    "condition_grade": "Surveyed condition A (as-new) .. D (unsuitable). Words like 'good'/'poor' map too.",
    "verification_status": "How the grade is trusted: mill_cert | coupon_tested | documented | "
                           "visual_only | unverified. Blank -> unverified (quarantined unless "
                           "--include-unverified).",
    "knockdown": "Optional explicit f_y knockdown (<=1). Blank -> derived from condition/verification.",
    "defects": "Free-text survey notes (corrosion, deformation, holes, ...).",
    "connection_type": "bolted | welded | riveted. Words like 'pinned'/'moment' map too.",
    "connection_condition": "Joint condition A..D.",
    "deconstructability": "easy | moderate | hard (override of the derived value).",
}

# Header (canonicalised: lowercased, non-alphanumerics stripped) -> our field. Distinct from the
# survey alias table on purpose: here a bare 'grade' means the STEEL grade, while in an audit survey
# it means the condition grade. Condition is reached via 'condition'/'state' instead.
_INVENTORY_ALIASES = {
    # identity
    "id": "id", "elementid": "id", "memberid": "id", "ref": "id", "piecenumber": "id", "piece": "id",
    "mark": "mark", "piecemark": "mark", "tag": "mark",
    "uniqueid": "unique_id", "guid": "unique_id", "globalid": "unique_id", "ifcguid": "unique_id",
    "role": "role", "memberrole": "role", "membertype": "role", "type": "role", "category": "role",
    "level": "level", "storey": "level", "story": "level", "floor": "level",
    "section": "raw_section", "rawsection": "raw_section", "sectiontype": "raw_section",
    "profile": "raw_section", "designation": "raw_section", "serialsize": "raw_section",
    "size": "raw_section",
    "grade": "material_grade", "materialgrade": "material_grade", "steelgrade": "material_grade",
    "material": "material_grade",
    "length": "length_mm", "lengthmm": "length_mm", "memberlength": "length_mm",
    "stocklength": "length_mm",
    "recoverablelength": "recoverable_length_mm", "recoverablelengthmm": "recoverable_length_mm",
    "usablelength": "recoverable_length_mm", "cutlength": "recoverable_length_mm",
    # audit / provenance
    "condition": "condition_grade", "conditiongrade": "condition_grade", "conditionstate": "condition_grade",
    "state": "condition_grade",
    "verification": "verification_status", "verificationstatus": "verification_status",
    "basis": "verification_status", "gradebasis": "verification_status", "provenance": "verification_status",
    "cert": "verification_status", "certification": "verification_status",
    "knockdown": "knockdown", "fyknockdown": "knockdown", "derate": "knockdown",
    "defects": "defects", "notes": "defects", "remarks": "defects",
    "connection": "connection_type", "connectiontype": "connection_type", "joint": "connection_type",
    "fixity": "connection_type",
    "connectioncondition": "connection_condition", "jointcondition": "connection_condition",
    "deconstructability": "deconstructability", "deconstruction": "deconstructability",
}

_NUMBER_FIELDS = ("length_mm", "recoverable_length_mm", "knockdown")
_AUDIT_FIELDS = ("condition_grade", "verification_status", "knockdown", "defects",
                 "connection_type", "connection_condition", "deconstructability")

# Common role synonyms -> canonical ROLES; unrecognised text falls back to "unknown".
_ROLE_SYNONYMS = {
    "beam": ("beam", "girder", "joist", "rafter", "purlin", "lintel", "spandrel"),
    "column": ("column", "post", "stanchion", "pile"),
    "brace": ("brace", "bracing", "diagonal", "strut", "tie"),
}


def _coerce_number(value):
    """Parse a numeric cell to float; None when blank/unparseable (so the schema default stands)."""
    if isinstance(value, (int, float)):
        return float(value)
    text = (value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_role(text):
    low = str(text or "").strip().lower()
    if not low:
        return "unknown"
    if low in ROLES:
        return low
    for canonical, words in _ROLE_SYNONYMS.items():
        if low in words:
            return canonical
    return "unknown"


def _map_header(name, col_map):
    if col_map and name in col_map:
        return col_map[name]
    return _INVENTORY_ALIASES.get(_canon_header(name))


def member_from_record(record):
    """Build an :class:`ExtractedMember` from a {field: raw_cell} record, or None for a blank row.

    ``id`` falls back to ``mark`` then ``unique_id``; a row with no identity *and* no section/length
    is treated as blank (skipped). Audit fields reuse the survey value-normalisation (synonyms +
    coercion); identity strings are kept verbatim, numbers coerced, role canonicalised.
    """
    get = lambda f: str(record.get(f) or "").strip()  # noqa: E731 -- local shorthand
    mid = get("id") or get("mark") or get("unique_id")
    section = get("raw_section")
    length = _coerce_number(record.get("length_mm"))
    if not mid and not section and length is None:
        return None  # entirely blank row

    fields = {
        "id": mid or "(unnamed)",
        "mark": get("mark") or None,
        "unique_id": get("unique_id") or None,
        "role": _coerce_role(record.get("role")),
        "level": get("level") or None,
        "raw_section": section,
        "material_grade": get("material_grade") or None,
        "length_mm": length if length is not None else 0.0,
    }
    rec_len = _coerce_number(record.get("recoverable_length_mm"))
    if rec_len is not None:
        fields["recoverable_length_mm"] = rec_len
    # Audit/provenance fields: set ONLY when the cell carries a value, so a bare id/section/length
    # inventory keeps the schema defaults (verification None -> admitted at the run knockdown) rather
    # than being coerced to "unverified" and silently quarantined. The conservative flag is opt-in,
    # via an explicit cell (as the template's example shows).
    for f in _AUDIT_FIELDS:
        raw = str(record.get(f) or "").strip()
        if not raw:
            continue
        value = normalize_survey_value(f, raw)
        if value is not None:
            fields[f] = value
    return ExtractedMember.from_dict(fields)


def load_model_file(path, kind):
    """Load an extraction model from JSON (the extractor's schema) or a spreadsheet inventory.

    A ``.csv``/``.tsv``/``.xlsx`` path is read by :func:`load_inventory_model` (headers mapped to the
    schema by alias); anything else goes through :meth:`ExtractedModel.load`. ``kind`` is stamped on a
    spreadsheet-built model ("donor" / "demand"). One dispatch point for every CLI entry that takes a
    model path, so spreadsheet inventories work in the matcher, the value case and review alike.
    """
    if Path(path).suffix.lower() in (".csv", ".tsv", ".xlsx", ".xlsm"):
        return load_inventory_model(str(path), kind=kind)
    return ExtractedModel.load(path)


def load_inventory_model(path, kind="donor", col_map=None, model_name=None):
    """Read a donor/demand inventory spreadsheet (.xlsx / .csv / .json) into an ``ExtractedModel``.

    Headers are mapped to schema fields by ``col_map`` (override) then the alias table; values are
    coerced/normalised. Blank rows are skipped. ``kind`` is "donor" or "demand". This is the engine
    behind both the CLI (``--donor inventory.xlsx``) and the pyRevit upload box.
    """
    members = []
    for raw_row in _read_rows(path):
        record = {}
        for src_header, cell in raw_row.items():
            field = _map_header(src_header, col_map)
            # First non-empty wins, so a stray duplicate/blank column cannot clobber a real value.
            if field is not None and not (record.get(field) or "").strip():
                record[field] = "" if cell is None else str(cell)
        member = member_from_record(record)
        if member is not None:
            members.append(member)
    return ExtractedModel(kind=kind, members=members, source="spreadsheet",
                          model_name=model_name or Path(path).stem)


# --- blank template authoring -------------------------------------------------------------------

def inventory_template_rows():
    """The template's data rows: a single worked example (headers come from INVENTORY_COLUMNS)."""
    return [{c: INVENTORY_EXAMPLE.get(c, "") for c in INVENTORY_COLUMNS}]


def inventory_template_csv():
    """Serialise the blank template (header + one example row) to CSV text."""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=INVENTORY_COLUMNS, lineterminator="\n")
    w.writeheader()
    for row in inventory_template_rows():
        w.writerow(row)
    return buf.getvalue()


def write_inventory_template_xlsx(path):
    """Write the blank template as a styled .xlsx: a data sheet (header + example) and a Guide sheet.

    Needs the optional ``openpyxl`` extra. The data sheet is the active/first sheet, so the reader
    picks it up unchanged; the example row is the same one the CSV template carries.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:  # noqa: BLE001
        raise RuntimeError("writing .xlsx needs the optional 'xlsx' extra: "
                           "pip install steelreuse[xlsx]") from e
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inventory"
    ws.append(INVENTORY_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in inventory_template_rows():
        ws.append([row.get(c, "") for c in INVENTORY_COLUMNS])

    guide = wb.create_sheet("Guide")
    guide.append(["column", "meaning"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
    for col in INVENTORY_COLUMNS:
        guide.append([col, COLUMN_HELP.get(col, "")])
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90
    wb.save(path)
    return path


def write_inventory_template(path):
    """Write a blank inventory template to ``path``; format chosen by extension (.xlsx else CSV)."""
    if Path(path).suffix.lower() in (".xlsx", ".xlsm"):
        return write_inventory_template_xlsx(path)
    Path(path).write_text(inventory_template_csv(), encoding="utf-8")
    return path
